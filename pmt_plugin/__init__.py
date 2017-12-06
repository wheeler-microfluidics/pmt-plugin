from collections import OrderedDict
import contextlib
import datetime as dt
import logging
import time
import serial
import warnings

from flatland import Boolean, Float, Form, Integer
from flatland.validation import ValueAtLeast, ValueAtMost
from microdrop.app_context import get_app
from microdrop.plugin_helpers import AppDataController, StepOptionsController
from microdrop.plugin_manager import (IPlugin, Plugin, implements, emit_signal,
                                      get_service_instance_by_name,
                                      PluginGlobals)
from mr_box_peripheral_board.max11210_adc_ui import MAX11210_begin

from pygtkhelpers.ui.objectlist import PropertyMapper
from pygtkhelpers.utils import dict_to_form
from pygtkhelpers.ui.extra_dialogs import yesno, FormViewDialog
import gobject
import gtk
import microdrop_utility as utility
import mr_box_peripheral_board as mrbox
import mr_box_peripheral_board.ui.gtk.measure_dialog
import openpyxl as ox
import openpyxl_helpers as oxh
import numpy as np
import pandas as pd
import path_helpers as ph
import inspect
import os

from ._version import get_versions
__version__ = get_versions()['version']
del get_versions


logger = logging.getLogger(__name__)


# Add plugin to `"microdrop.managed"` plugin namespace.
PluginGlobals.push_env('microdrop.managed')


# Path to DRC data collection template Excel spreadsheet workbook.
TEMPLATE_PATH = (ph.path(r'templates')
                 .joinpath('DRC Data Collection-named_ranges.xlsx'))

dirname = os.path.dirname(os.path.abspath(inspect.stack()[0][1]))

def _write_results(template_path, output_path, data_files):
    '''
    Write results as Excel spreadsheet to output path based on template.

    .. versionadded:: 0.19

    Parameters
    ----------
    template_path : str
        Path to Excel template spreadsheet.
    output_path : str
        Path to write output Excel spreadsheet to.
    data_files : list
        List of paths to `new-line delimited JSON files <http://ndjson.org/`_
        containing measured PMT data.

        Each new-line delimited JSON file corresponds to protocol step, and
        each line in each file corresponds to measured PMT data from a single
        measurement run.

    Returns
    -------
    path_helpers.path
        Wrapped output path.

        Allows, for example, easy opening of document using the ``launch()``
        method.
    '''
    output_path = ph.path(output_path)

    logger.info('---IN FUNC---')

    # XXX `openpyxl` does not currently [support reading existing data
    # validation][1].
    #
    # As a workaround, load the extension lists and data validation definitions
    # from the template workbook so they may be restored to the output workbook
    # after modifying with `openpyxl`.
    #
    # [1]: http://openpyxl.readthedocs.io/en/default/validation.html#validating-cells
    #template_path = r"C:\Users\Queen's Park\MicroDrop\etc\MicroDrop\plugins\enabled\mr_box_plugin_Legacy\templates\DRC Data Collection-named_ranges.xlsx"
    #template_path = r"templates\DRC Data Collection-named_ranges.xlsx"
    #C:\Users\Queen's Park\Documents\Microdrop
    template_path = os.path.join(dirname, TEMPLATE_PATH)
    logger.info("---TEMPLATE PATH---")
    logger.info(template_path)

    extension_lists = oxh.load_extension_lists(template_path)
    data_validations = oxh.load_data_validations(template_path)

    logger.info('---LOADED---')

    with warnings.catch_warnings():
        warnings.filterwarnings('ignore', 'Data Validation extension is not '
                                'supported and will be removed', UserWarning)

        # Open template workbook to modify it in-memory before writing to the
        # output file.
        with contextlib.closing(ox.load_workbook(template_path)) as workbook:
            # Create pandas Excel writer to make it easier to append data
            # frames as worksheets.
            with pd.ExcelWriter(output_path,
                                engine='openpyxl') as output_writer:
                # Configure pandas Excel writer to append to template workbook
                # contents.
                output_writer.book = workbook

                # Get mapping from each worksheet name to the correpsonding
                # worksheet object.
                worksheets_by_name = OrderedDict(zip(workbook.sheetnames,
                                                     workbook.worksheets))

                # Get list of defined names, grouped by worksheet name.
                defined_names_by_worksheet = \
                    oxh.get_defined_names_by_worksheet(workbook)

                # Select the "Location" entry cell by default.
                workbook.active = 0
                worksheet = worksheets_by_name['Assay Info']
                default_cell = \
                    worksheet[defined_names_by_worksheet['Assay Info']
                              ['LocationEntry']]
                for attribute_i in ('activeCell', 'sqref'):
                    setattr(worksheet.views.sheetView[0].selection[0],
                            attribute_i, default_cell.coordinate)

                # Set the "Date" entry cell value to the current date.
                date_cell = \
                    worksheet[defined_names_by_worksheet['Assay Info']
                              ['DateEntry']]
                date_cell.value = dt.datetime.utcnow().date()

                # Look up the location of the top cell in the measurement IDs
                # column of the PMT results information table.
                pmt_ids_range = (defined_names_by_worksheet['Assay Info']
                                 ['PMTMeasurementIDEntries'])
                pmt_ids_boundaries = ox.utils.range_boundaries(pmt_ids_range)
                pmt_ids_column = pmt_ids_boundaries[0]

                # Look up the location of the top cell in the PMT mean
                # measurements column of the results information table.
                pmt_mean_range = (defined_names_by_worksheet['Assay Info']
                                  ['PMTMeanEntries'])
                pmt_mean_boundaries = ox.utils.range_boundaries(pmt_mean_range)
                pmt_mean_column = pmt_mean_boundaries[0]

                # Set output row index to the first row of the PMT results
                # table.
                pmt_results_row = pmt_ids_boundaries[1]

                # Write the data from each PMT measurement run to a separate:
                #  1. **worksheet**; and
                #  2. **row** in the PMT results information table in the
                #     `Assay Info` worksheet.
                for data_file_i in data_files:
                    # Each line in each [new-line delimited JSON file][1]
                    # corresponds to measured PMT data from a single
                    # measurement run.
                    #
                    # [1]: http://ndjson.org/
                    for j, data_json_ij in enumerate(data_file_i.lines()):
                        try:
                            # Try reading JSON data with `split` orientation,
                            # which preserves the name of the Pandas series.
                            s_data_ij = pd.read_json(data_json_ij,
                                                     typ='series',
                                                     orient='split')
                        except ValueError:
                            logging.debug('Decode legacy series')
                            # JSON data was not encoded in `split` orientation.
                            s_data_ij = pd.read_json(data_json_ij,
                                                     typ='series')

                        if not s_data_ij.name:
                            # No name was encoded in the JSON data.  Interpret
                            # data series name from filename.
                            s_data_ij.name = '%s-%02d' % (data_file_i.namebase
                                                          .split('-')[-1], j)

                        # Add column indicating time of each sample relative to
                        # time of first sample point for easier comparison
                        # between worksheets.
                        relative_time_s_i = (pd.Series(s_data_ij.index -
                                                       s_data_ij.index[0])
                                             .dt.total_seconds())
                        df_data_ij = s_data_ij.to_frame()
                        df_data_ij.insert(0, 'relative_time_s',
                                          relative_time_s_i.values)

                        # Write measurement data to worksheet.
                        df_data_ij.to_excel(output_writer,
                                            sheet_name=s_data_ij.name,
                                            header=True)

                        # Write name of measurement run to the PMT results
                        # information table in the `Assay Info` worksheet.
                        id_cell_ij = worksheet.cell(row=pmt_results_row,
                                                    column=pmt_ids_column)
                        id_cell_ij.value = s_data_ij.name

                        # Write formula for the average measurement value to
                        # the PMT results information table in the `Assay Info`
                        # worksheet.
                        mean_cell_ij = worksheet.cell(row=pmt_results_row,
                                                      column=pmt_mean_column)

                        sheetname_ij = ox.utils.quote_sheetname(s_data_ij.name)
                        mean_formula_ij = ('=AVERAGE({sheetname}!C2, '
                                           '{sheetname}!C{end_row})'
                                           .format(sheetname=sheetname_ij,
                                                   end_row=2 + len(s_data_ij)))
                        mean_cell_ij.value = mean_formula_ij

                        # Set output row index to the next row of the PMT
                        # results table.
                        pmt_results_row += 1

                # Create mapping of the name of each worksheet containing PMT
                # measurement data to the corresponding worksheet.
                chart_worksheets_by_name = OrderedDict(zip(workbook
                                                           .sheetnames[2:],
                                                           workbook
                                                           .worksheets[2:]))

                # Generate list of colors to use for plotting PMT measurements.
                # Randomize order (deterministically due to static seed) since
                # default order is alphabetic.
                random_state = np.random.RandomState(1)
                colors = ox.drawing.colors.PRESET_COLORS[:]
                random_state.shuffle(colors)

                # Create global chart for plotting all measurements to the
                # `Assay Info` worksheet.
                chart = ox.chart.ScatterChart()
                chart.x_axis.title = 'Time (s)'
                chart.y_axis.title = 'Current (A)'

                for i, (name_i,
                        worksheet_i) in enumerate(chart_worksheets_by_name
                                                  .iteritems()):
                    # Create chart for current PMT measurements worksheet.
                    chart_i = ox.chart.ScatterChart()
                    chart_i.x_axis.title = chart.x_axis.title
                    chart_i.y_axis.title = chart.y_axis.title

                    # Find bottom row index in current worksheet.
                    max_row_i = max([cell_i.row for cell_i in
                                     worksheet.get_cell_collection()])

                    # Select color for current worksheet.
                    color_i = ox.drawing.colors.ColorChoice(prstClr=colors
                                                            [i % len(colors)])

                    # Create data series referring to PMT data from current
                    # worksheet.
                    xvalues_i = ox.chart.Reference(worksheet_i, min_col=2,
                                                   min_row=2,
                                                   max_row=max_row_i)
                    yvalues_i = ox.chart.Reference(worksheet_i, min_col=3,
                                                   min_row=1,
                                                   max_row=max_row_i)
                    series_i = ox.chart.Series(yvalues_i, xvalues_i,
                                               title_from_data=True)
                    # Set line color for measurements from current worksheet.
                    line_prop_i = \
                        ox.drawing.line.LineProperties(solidFill=color_i)
                    series_i.graphicalProperties.line = line_prop_i
                    chart_i.title = 'PMT current'
                    chart_i.height = 10  # default is 7.5
                    chart_i.width = 20  # default is 15

                    # Add chart to current worksheet.
                    chart_i.series.append(series_i)

                    worksheet_i.add_chart(chart_i, 'D1')

                    # Add PMT data series from current worksheet to common
                    # chart in `Assay Info` worksheet.
                    chart.series.append(series_i)

                # Add common chart containing data from all PMT worksheets to
                # `Assay Info` worksheet.
                chart.title = 'PMT current'
                chart.height = 20  # default is 7.5
                chart.width = 25  # default is 15

                worksheet.add_chart(chart, 'I1')

    # Restore the extension lists and data validation definitions to the output
    # workbook (they were removed by `openpyxl`, see above).
    logger.info('---REACHED END---')
    updated_xlsx = oxh.update_extension_lists(output_path, extension_lists)
    with output_path.open('wb') as output:
        output.write(updated_xlsx)
    updated_xlsx = oxh.update_data_validations(output_path, data_validations)
    with output_path.open('wb') as output:
        output.write(updated_xlsx)

    return output_path


class MrBoxPeripheralBoardPlugin(AppDataController, StepOptionsController,
                                 Plugin):
    '''
    This class is automatically registered with the PluginManager.
    '''
    implements(IPlugin)

    plugin_name = str(ph.path(__file__).realpath().parent.name)
    try:
        version = __version__
    except NameError:
        version = 'v0.0.0+unknown'

    AppFields = Form.of(Boolean.named('Use PMT y-axis SI units')
                        .using(default=True, optional=True))

    StepFields = Form.of(# PMT Fields
                         Boolean.named('Measure_PMT')
                         .using(default=False, optional=True),
                         # Only allow PMT Duration to be set if `Measure_PMT`
                         # field is set to `True`.
                         Integer.named('Measurement_duration_(s)')
                         .using(default=10, optional=True,
                                validators=[ValueAtLeast(minimum=0)],
                                properties={'mappers':
                                            [PropertyMapper
                                             ('sensitive', attr='Measure_PMT'),
                                             PropertyMapper
                                             ('editable',
                                              attr='Measure_PMT')]}))
                         # Only allow ADC Gain to be set if `Measure_PMT` field
                         # is set to `True`.
                         # TODO Convert ADC Gain to dropdown list with
                         # valid_values = (1,2,4,8,16)
                        #  Integer.named('ADC_Gain')
                        #  .using(default=1, optional=True,
                        #         validators=[ValueAtLeast(minimum=1),
                        #                     ValueAtMost(maximum=16)],
                        #         properties={'mappers':
                        #                     [PropertyMapper
                        #                      ('sensitive', attr='Measure_PMT'),
                        #                      PropertyMapper
                        #                      ('editable',
                        #                       attr='Measure_PMT')]}),


    def __init__(self):
        super(MrBoxPeripheralBoardPlugin, self).__init__()
        self.board = None
        # XXX `name` attribute is required in addition to `plugin_name`
        #
        # The `name` attribute is required in addition to the `plugin_name`
        # attribute because MicroDrop uses it for plugin labels in, for
        # example, the plugin manager dialog.
        self.name = self.plugin_name

        # Flag to indicate whether user has already been warned about the board
        # not being connected when trying to set board state.
        self._user_warned = False

        # `dropbot.SerialProxy` instance
        self.dropbot_remote = None

        # Latch to, e.g., config menus, only once
        self.initialized = False

        self.adc_gain_calibration = None
        self.adc_offset_calibration = None
        self.off_cal_val = None

    def reset_board_state(self):
        '''
        Reset MR-Box peripheral board to default state.
        '''
        # Reset user warned state (i.e., warn user next time board settings
        # are applied when board is not connected).
        self._user_warned = False

        if self.board is None:
            return

        # TODO Add reset method for each component (PMT)
        # TODO to respective `mr-box-peripheral-board.py` C++ classes code.






        # Set PMT control voltage to zero.
        self.board.pmt_set_pot(0)
        # Start the ADC and Perform ADC Calibration
        MAX11210_begin(self.board)



    def apply_step_options(self, step_options):
        '''
        Apply the specified step options.

        Parameters
        ----------
        step_options : dict
            Dictionary containing the MR-Box peripheral board plugin options
            for a protocol step.
        '''

        app = get_app()
        app_values = self.get_app_values()

        if self.board:
            step_log = {}

            services_by_name = {service_i.name: service_i
                                for service_i in
                                PluginGlobals
                                .env('microdrop.managed').services}

            step_label = None
            if 'wheelerlab.step_label_plugin' in services_by_name:
                # Step label is set for current step
                step_label_plugin = (services_by_name
                                     .get('wheelerlab.step_label_plugin'))
                step_label = (step_label_plugin.get_step_options()
                              or {}).get('label')

            # Apply board hardware options.
            try:

                # PMT/ADC
                # -------
                if step_options.get('Measure_PMT'):

                    # Start the ADC and Perform ADC Calibration
                    MAX11210_begin(self.board)

                    if step_label.lower() == 'background':
                        ''' Set PMT control voltage via digipot.'''
                        # Divide the control voltage by the maximum 1100 mV and
                        # convert it to digipot steps
                        '''
                        Perform certain calibration steps only for the background
                        measurement.

                        Read from the 24bit Registries (SCGC, SCOC)
                        and store their values for the rest of the
                        measurements.
                        '''

                        logger.warning('Open PMT shutter and close box lid')
                        self.adc_gain_calibration = self.board.MAX11210_getSelfCalGain()
                        self.adc_offset_calibration = self.board.MAX11210_getSelfCalOffset()
                        self.board.MAX11210_setSysOffsetCal(0x00)
                        self.board.MAX11210_send_command(0b10001000)
                        reading_i = []
                        for i in range(0,20):
                            self.board.MAX11210_setRate(120)
                            reading_i.append(self.board.MAX11210_getData())
                        reading_avg = (sum(reading_i)* 1.0) / (len(reading_i) * 1.0)
                        self.off_cal_val = int(reading_avg) - 1677

                    else:
                        if not self.adc_gain_calibration:
                            logger.warning('Missing ADC Calibration Values!'
                                            'Please perform a Background measurement')
                        else:
                            logger.warning('Open PMT shutter and close box lid')
                            self.board.MAX11210_setSelfCalGain(self.adc_gain_calibration)
                            self.board.MAX11210_setSelfCalOffset(self.adc_offset_calibration)
                    '''if (self.board.config.pmt_sys_offset_cal != 0):
                        self.board.MAX11210_setSysOffsetCal(self.board.config.pmt_sys_offset_cal)
                    else:
                        self.board.MAX11210_setSysOffsetCal(self.off_cal_val)
                    self.board.MAX11210_setSysGainCal(self.board.config.pmt_sys_gain_cal)
                    self.board.MAX11210_send_command(0b10001000)'''

                    adc_calibration = self.board.get_adc_calibration().to_dict()
                    logger.info('ADC calibration:\n%s' % adc_calibration)
                    step_log['ADC calibration'] = adc_calibration


                    # Launch PMT measure dialog.
                    delta_t = dt.timedelta(seconds=1)

                    # Set sampling reset_board_state
                    adc_rate = self.board.config.pmt_sampling_rate
                    # Construct a function compatible with `measure_dialog` to
                    # read from MAX11210 ADC.
                    data_func = (mrbox.ui.gtk.measure_dialog
                                 .adc_data_func_factory(proxy=self.board,
                                                        delta_t=delta_t,
                                                        adc_rate=adc_rate))

                    # Use constructed function to launch measurement dialog for
                    # the duration specified by the step options.
                    duration_s = (step_options.get('Measurement_duration_(s)')
                                  + 1)
                    use_si_prefixes = app_values.get('Use PMT y-axis SI '
                                                     'prefixes')
                    data = (mrbox.ui.gtk.measure_dialog
                            .measure_dialog(data_func, duration_s=duration_s,
                                            auto_start=True, auto_close=False,
                                            si_units=use_si_prefixes))
                    if data is not None:
                        # Append measured data as JSON line to [new-line
                        # delimited JSON][1] file for step.
                        #
                        # Each line of results can be loaded using
                        # `pandas.read_json(..., orient='split')`.
                        #
                        # [1]: http://ndjson.org/
                        filename = ph.path('PMT_readings-step%04d.ndjson' %
                                    app.protocol.current_step_number)
                        log_dir = app.experiment_log.get_log_path()
                        log_dir.makedirs_p()

                        data.name = filename.namebase

                        if step_label:
                            # Set name of data series based on step label.
                            data.name = step_label

                        with log_dir.joinpath(filename).open('a') as output:
                            # Write JSON data with `split` orientation, which
                            # preserves the name of the Pandas series.
                            data.to_json(output, orient='split')
                            output.write('\n')

                        step_log['data'] = data.to_dict()

                        self.update_excel_results()
                        logger.warning('Close PMT Shutter')
            except Exception:
                logger.error('[%s] Error applying step options.', __name__,
                             exc_info=True)
            finally:
                app.experiment_log.add_data(step_log, self.name)

        elif not self._user_warned:
            logger.warning('[%s] Cannot apply board settings since board is '
                           'not connected.', __name__, exc_info=True)
            # Do not warn user again until after the next connection attempt.
            self._user_warned = True

    def update_excel_results(self, launch=False):
        '''
        Update output Excel results file.

        .. versionadded:: 0.19

        Parameters
        ----------
        launch : bool, optional
            If ``True``, launch Excel spreadsheet after writing.
        '''
        app = get_app()
        log_dir = app.experiment_log.get_log_path()

        # Update Excel file with latest PMT results.
        output_path = log_dir.joinpath('PMT_readings.xlsx')
        data_files = list(log_dir.files('PMT_readings-*.ndjson'))

        if not data_files:
            logger.debug('No PMT readings files found.')
            return

        logger.info(TEMPLATE_PATH)

        def _threadsafe_write_results():
            logger.info(launch)
            while True:
                try:
                    _write_results(TEMPLATE_PATH, output_path, data_files)
                    if launch:
                        try:
                            output_path.launch()
                        except Exception:
                            pass
                    break
                except IOError as e:
                    logger.info("I/O error({0}): {1}".format(e.errno, e.strerror))
                    response = yesno('Error writing PMT summary to Excel '
                                     'spreadsheet output path: `%s`.\n\nTry '
                                     'again?' %output_path)
                    if response == gtk.RESPONSE_NO:
                        break

        # Schedule writing of results to occur in main GTK
        # thread in case confirmation dialog needs to be
        # displayed.
        gobject.idle_add(_threadsafe_write_results)



    def open_board_connection(self):
        '''
        Establish serial connection to MR-Box peripheral board.
        '''
        # Try to connect to peripheral board through serial connection.

        # XXX Try to connect multiple times.
        # See [issue 1][1] on the [MR-Box peripheral board firmware
        # project][2].
        #
        # [1]: https://github.com/wheeler-microfluidics/mr-box-peripheral-board.py/issues/1
        # [2]: https://github.com/wheeler-microfluidics/mr-box-peripheral-board.py
        retry_count = 2
        for i in xrange(retry_count):
            try:
                self.board.close()
                self.board = None
            except Exception:
                pass

            try:

                self.board = mrbox.SerialProxy()

                host_software_version = utility.Version.fromstring(
                    str(self.board.host_software_version))
                remote_software_version = utility.Version.fromstring(
                    str(self.board.remote_software_version))

                # Offer to reflash the firmware if the major and minor versions
                # are not not identical. If micro versions are different,
                # the firmware is assumed to be compatible. See [1]
                #
                # [1]: https://github.com/wheeler-microfluidics/base-node-rpc/
                #              issues/8
                if any([host_software_version.major !=
                        remote_software_version.major,
                        host_software_version.minor !=
                        remote_software_version.minor]):
                    response = yesno("The MR-box peripheral board firmware "
                                     "version (%s) does not match the driver "
                                     "version (%s). Update firmware?" %
                                     (remote_software_version,
                                      host_software_version))
                    if response == gtk.RESPONSE_YES:
                        self.on_flash_firmware()

                # Serial connection to peripheral **successfully established**.
                logger.info('Serial connection to peripheral board '
                            '**successfully established**')

                logger.info('Peripheral board properties:\n%s',
                            self.board.properties)
                logger.info('Reset board state to defaults.')
                break
            except (serial.SerialException, IOError):
                time.sleep(1)
        else:
            # Serial connection to peripheral **could not be established**.
            logger.warning('Serial connection to peripheral board could not '
                           'be established.')

    def on_edit_configuration(self, widget=None, data=None):
        '''
        Display a dialog to manually edit the configuration settings.
        '''
        config = self.board.config
        form = dict_to_form(config)
        dialog = FormViewDialog(form, 'Edit configuration settings')
        valid, response = dialog.run()
        if valid:
            self.board.update_config(**response)

    def on_flash_firmware(self, widget=None, data=None):
        app = get_app()
        try:
            self.board.flash_firmware()
            app.main_window_controller.info("Firmware updated successfully.",
                                            "Firmware update")
        except Exception, why:
            logger.error("Problem flashing firmware. ""%s" % why)

    def close_board_connection(self):
        '''
        Close serial connection to MR-Box peripheral board.
        '''
        if self.board is not None:
            # Close board connection and release serial connection.
            self.board.close()

    def get_schedule_requests(self, function_name):
        """
        Returns a list of scheduling requests (i.e., ScheduleRequest
        instances) for the function specified by function_name.
        """
        # TODO: this should be re-enabled once we can get the
        # mr-box-peripheral-board to connect **after** the Dropbot.
        # if function_name in ['on_plugin_enable']:
        #    return [ScheduleRequest('dropbot_plugin', self.name)]
        return []

    ###########################################################################
    # MicroDrop pyutillib plugin signal handlers
    # ------------------------------------------
    ###########################################################################
    def on_plugin_enable(self):
        '''
        Handler called when plugin is enabled.

        For example, when the MicroDrop application is **launched**, or when
        the plugin is **enabled** from the plugin manager dialog.
        '''
        self.open_board_connection()
        if not self.initialized:
            app = get_app()
            self.tools_menu_item = gtk.MenuItem("PMT-plugin")
            app.main_window_controller.menu_tools.append(self.tools_menu_item)
            self.tools_menu = gtk.Menu()
            self.tools_menu_item.set_submenu(self.tools_menu)

            self.edit_config_menu_item = \
                gtk.MenuItem("Edit configuration settings...")
            self.tools_menu.append(self.edit_config_menu_item)
            self.edit_config_menu_item.connect("activate",
                                               self.on_edit_configuration)
            self.edit_config_menu_item.show()
            self.initialized = True

        # if we're connected to the board, display the menu
        if self.board:
            self.reset_board_state()
            self.tools_menu.show()
            self.tools_menu_item.show()

        try:
            super(MrBoxPeripheralBoardPlugin, self).on_plugin_enable()
        except AttributeError:
            pass

    def initialize_connection_with_dropbot(self):
        '''
        If the dropbot plugin is installed and enabled, try getting its
        reference.
        '''
        try:
            service = get_service_instance_by_name('dropbot_plugin')
            if service.enabled():
                self.dropbot_remote = service.control_board
            assert(self.dropbot_remote.properties.package_name == 'dropbot')
        except Exception:
            logger.debug('[%s] Could not communicate with Dropbot.', __name__,
                         exc_info=True)
            logger.warning('Could not communicate with DropBot.')

    def on_plugin_disable(self):
        '''
        Handler called when plugin is disabled.

        For example, when the MicroDrop application is **closed**, or when the
        plugin is **disabled** from the plugin manager dialog.
        '''
        try:
            super(MrBoxPeripheralBoardPlugin, self).on_plugin_disable()
        except AttributeError:
            pass
        self.close_board_connection()
        self.tools_menu.hide()
        self.tools_menu_item.hide()

    def on_protocol_run(self):
        '''
        Handler called when a protocol starts running.
        '''
        # TODO: this should be run in on_plugin_enable; however, the
        # mr-box-peripheral-board seems to have trouble connecting **after**
        # the DropBot has connected.
        self.initialize_connection_with_dropbot()


    def on_protocol_pause(self):
        # Protocol has finished.  Update
        app = get_app()
        if ((app.protocol.current_step_number == len(app.protocol) - 1) and
            (app.protocol.current_repetition == app.protocol.n_repeats - 1)):
            self.update_excel_results(launch=True) #JOCHANGE

    def on_experiment_log_changed(self, experiment_log):
        '''
        Handler called when a new experiment starts.
        '''
        logger.info('Reset board state to defaults.')
        if self.board:
            self.reset_board_state()


    def on_app_options_changed(self, plugin_name):
        """
        Handler called when the app options are changed for a particular
        plugin.  This will, for example, allow for GUI elements to be
        updated.

        Parameters
        ----------
        plugin : str
            Plugin name for which the app options changed
        """


    def on_step_options_changed(self, plugin, step_number):
        '''
        Handler called when field values for the specified plugin and step.

        Parameters
        ----------
        plugin : str
            Name of plugin.
        step_number : int
            Step index number.
        '''
        # Step options have changed.
        app = get_app()

        if all([plugin == self.plugin_name, app.realtime_mode,
                step_number == app.protocol.current_step_number]):
            # Apply step options.
            options = self.get_step_options()
            self.apply_step_options(options)

    def on_step_run(self):
        '''
        Handler called whenever a step is executed.

        Plugins that handle this signal **MUST** emit the ``on_step_complete``
        signal once they have completed the step.  The protocol controller will
        wait until all plugins have completed the current step before
        proceeding.
        '''
        # Get latest step field values for this plugin.
        options = self.get_step_options()
        # Apply step options
        self.apply_step_options(options)

        emit_signal('on_step_complete', [self.name])

    def on_step_swapped(self, original_step_number, new_step_number):
        '''
        Handler called when a new step is activated/selected.

        Parameters
        ----------
        original_step_number : int
            Step number of previously activated step.
        new_step_number : int
            Step number of newly activated step.
        '''
        # Step options have changed.
        app = get_app()
        if app.realtime_mode and not app.running:
            # Apply step options.
            options = self.get_step_options()
            self.apply_step_options(options)


PluginGlobals.pop_env()
